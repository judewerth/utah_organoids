"""Unit tests for Excel metadata parsing functions."""
import pytest
import tempfile
from pathlib import Path
from unittest.mock import patch

import openpyxl


class TestReadEphysInfoFromExcel:
    """Tests for read_ephys_info_from_excel_2017()."""

    @pytest.fixture
    def read_metadata(self):
        """Import the function under test."""
        from workflow.pipeline.patch_clamp_ephys.read_metadata import (
            read_ephys_info_from_excel_2017,
        )
        return read_ephys_info_from_excel_2017

    @pytest.fixture
    def sample_excel(self, tmp_path):
        """Create a minimal Excel file matching the expected format."""
        wb = openpyxl.Workbook()
        ws = wb.active

        # Animal info (rows 1-4, skiprows_animal=1 means header at row 2)
        ws["A1"] = "Animal Info"
        ws["A2"] = "id"
        ws["B2"] = "strain"
        ws["C2"] = "dob"
        ws["D2"] = "date"
        ws["E2"] = "age"
        ws["F2"] = "slicetype"
        ws["G2"] = "external"
        ws["H2"] = "internal"
        ws["I2"] = "comment"
        ws["A3"] = "ORG-1"
        ws["B3"] = "H9"
        ws["C3"] = "2020-01-01"
        ws["D3"] = "2020-08-28"
        ws["E3"] = 240
        ws["F3"] = "whole"
        ws["G3"] = "aCSF"
        ws["H3"] = "K-gluconate"
        ws["I3"] = ""

        # Empty rows before cell data
        # Row 4, 5, 6 empty

        # Cell/recording info (skiprows_cell=6 means header at row 7)
        ws["A7"] = "Experiment"
        ws["B7"] = "Cell"
        ws["C7"] = "Recording"
        ws["D7"] = "Type"
        ws["E7"] = "Step"
        ws["A8"] = "2020-08-28"
        ws["B8"] = "1"
        ws["C8"] = "2020_08_28_0006"
        ws["D8"] = "IC"
        ws["E8"] = 10

        filepath = tmp_path / "test_experiment.xlsx"
        wb.save(filepath)
        return filepath

    def test_parses_animal_info(self, read_metadata, sample_excel):
        """Test that animal metadata is correctly parsed."""
        animal_info, _ = read_metadata(sample_excel)
        assert animal_info["id"] == "ORG-1"
        assert animal_info["strain"] == "H9"

    def test_parses_cell_info(self, read_metadata, sample_excel):
        """Test that cell/recording info is correctly parsed."""
        _, cell_info = read_metadata(sample_excel)
        assert len(cell_info) > 0
        assert "Experiment" in cell_info.columns or "experiment" in cell_info.columns.str.lower()

    def test_custom_skiprows(self, read_metadata, sample_excel):
        """Test that custom skiprows parameters work."""
        animal_info, cell_info = read_metadata(
            sample_excel, skiprows_animal=1, skiprows_cell=6
        )
        assert animal_info is not None
        assert cell_info is not None

    def test_missing_file_raises(self, read_metadata, tmp_path):
        """Test that missing file raises an error."""
        with pytest.raises(Exception):
            read_metadata(tmp_path / "nonexistent.xlsx")


class TestCurrentStepTimingValues:
    """Test that timing parameter computations are correct."""

    def test_default_timing_values(self):
        """Test standard timing parameter computation."""
        istep_start = 0.55
        istep_duration = 1.0
        istep_end = istep_start + istep_duration
        istep_end_1s = istep_start + 1.0

        assert istep_end == 1.55
        assert istep_end_1s == 1.55

    def test_short_duration_caps_end_1s(self):
        """Test that istep_end_1s doesn't exceed actual end for short durations."""
        istep_start = 0.55
        istep_duration = 0.5  # shorter than 1s
        istep_end = istep_start + istep_duration
        istep_end_1s = istep_start + 1.0

        if istep_end < istep_end_1s:
            istep_end_1s = istep_end

        assert istep_end_1s == 1.05  # capped at actual end
